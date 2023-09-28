import os
from typing import Any

import openpyxl
from io import BytesIO

import pandas as pd
from pandas import DataFrame

from modules.excel_handler import ExcelHandler


def read_page_line():

    status = "Error"
    # read article count in storage
        # read article id
            # create product
            # read other atribbutes
                #edit atributes

    status = "Complete"
    return status


class OtomotoManager:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.excel_handler = ExcelHandler()


    def _create_page(self):
        self.file_name = r"Test_doc.xlsx"
        self.sheet_name = "OtoMoto"
        print(self.file_name)
        print(self.sheet_name)
        file_content = self.excel_handler.get_exel_file(self.file_name)


        current_directory = os.getcwd()
        save_path = os.path.join(current_directory, 'Data', self.file_name)

        with open(save_path, 'wb') as file:
            file.write(file_content)


        #
        # wb = openpyxl.load_workbook('D:\API-Spaceplus\Data\Test_doc.xlsx')
        # sheet = wb.active

        # for row in sheet.iter_rows():
        #     for cell in row:
        #         print(cell.value)

        # 1 variant
        # qqwe = openpyxl.load_workbook('D:\API-Spaceplus\Data\Test_doc.xlsx')
        # sheet = qqwe.active
        # for row in sheet.iter_rows():
        #     for cell in row:
        #         print(cell.value)

        # 2 variant
        file_path = save_path
        df1 = pd.read_excel(file_path)
        for column in df1.columns:
            print(f"Column: {column}")
            for value in df1[column]:
                print(value)

        return self

    def read_all_items(self):
        self._create_page()

        pass
